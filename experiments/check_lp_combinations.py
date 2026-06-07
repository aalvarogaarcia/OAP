"""experiments/check_lp_combinations.py
Evaluate LP-relaxation quality across combinations of three constraint families:

- crossing constraints (version 1 only): off/on
- bipartition constraints: mode 0/1/2
- triangle crossing constraints: mode 0/1/2

For each instance and direction (MIN/MAX), solve the relaxed MCF model for all
18 combinations, rank them by LP bound quality, and export:

1) Excel workbook with visual comparison (full, winners, leaderboard)
2) Optional JSON report with violated inequalities for worst combinations

Run from repo root:
    c:/Users/ALVARO/Documents/codigo/OAP/.venv/Scripts/python.exe experiments/check_lp_combinations.py
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Literal, cast

_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT))

import pandas as pd  # noqa: E402
from openpyxl.formatting.rule import DataBarRule  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from models import OAPCompactModel  # noqa: E402
from utils.utils import (  # noqa: E402
    build_crossing_arc_index,
    compute_crossing_edges,
    compute_triangles,
    iter_directed_crossing_pairs,
    read_indexed_instance,
)

INSTANCE_DIR: Path = _REPO_ROOT / "instance"
TSV_PATH: Path = _REPO_ROOT / "test" / "data" / "TablaResultadosA4.tsv"
OUTPUT_XLSX: Path = _REPO_ROOT / "outputs" / "CSV" / "lp_combo_comparison.xlsx"
OUTPUT_JSON: Path = _REPO_ROOT / "outputs" / "JSON" / "lp_combo_violations.json"

OBJECTIVE: Literal["Fekete", "Internal", "External", "Diagonals"] = "Internal"
MODE: int = 0
SUM_CONSTRAIN: bool = True
STRENGTHEN: bool = False
TIME_LIMIT: int = 300
VIOL_TOL: float = 1e-6


@dataclass(frozen=True)
class Combo:
    crossing_on: bool
    bip_mode: int
    tri_mode: int

    @property
    def combo_id(self) -> str:
        return f"c{int(self.crossing_on)}_b{self.bip_mode}_t{self.tri_mode}"


def _all_combos() -> list[Combo]:
    combos: list[Combo] = []
    for crossing_on in (False, True):
        for bip_mode in (0, 1, 2):
            for tri_mode in (0, 1, 2):
                combos.append(Combo(crossing_on=crossing_on, bip_mode=bip_mode, tri_mode=tri_mode))
    return combos


def _numeric(v: float | str) -> bool:
    return isinstance(v, (int, float))


def _bound_delta(direction: str, lp: float | str, baseline: float | str) -> float | str:
    if not (_numeric(lp) and _numeric(baseline)):
        return "-"
    if direction == "MIN":
        return float(lp) - float(baseline)
    return float(baseline) - float(lp)


def _solve_relaxed(instance_file: Path, maximize: bool, combo: Combo, time_limit: int) -> tuple[float | str, OAPCompactModel]:
    direction = "max" if maximize else "min"
    points = read_indexed_instance(str(instance_file))
    triangles = compute_triangles(points)
    name = f"{instance_file.stem}_{direction}_{combo.combo_id}"

    model = OAPCompactModel(points, triangles, name=name)
    model.build(
        objective=OBJECTIVE,
        mode=MODE,
        maximize=maximize,
        subtour="MCF",
        sum_constrain=SUM_CONSTRAIN,
        strengthen=STRENGTHEN,
    )

    if combo.crossing_on:
        model._add_crossing_constraints(version=1)
    model._add_bipartition_constraints(mode=combo.bip_mode)
    model._add_triangle_crossing_constraints(mode=combo.tri_mode)

    model.solve(time_limit=time_limit, relaxed=True, verbose=False)
    return model.get_objval_lp(), model


def _sort_key(direction: str, lp: float | str) -> float:
    if not _numeric(lp):
        return float("inf")
    value = float(lp)
    return -value if direction == "MIN" else value


def _crossing_violations(model: OAPCompactModel, tol: float) -> list[dict[str, Any]]:
    violations: list[dict[str, Any]] = []
    crossing = compute_crossing_edges(model.triangles, model.points)

    for (u, v), (w, z) in iter_directed_crossing_pairs(crossing):
        if (u, v) not in model.x or (w, z) not in model.x:
            continue
        x_uv = model.x_relaxed.get((u, v), 0.0)
        x_wz = model.x_relaxed.get((w, z), 0.0)
        lhs = x_uv + x_wz
        rhs = 1.0
        viol = lhs - rhs
        if viol > tol:
            violations.append(
                {
                    "family": "crossing_v1",
                    "mode": "pairwise",
                    "key_indices": {"u": u, "v": v, "w": w, "z": z},
                    "lhs": lhs,
                    "rhs": rhs,
                    "violation": viol,
                    "top_terms": [
                        {"arc": [u, v], "value": x_uv},
                        {"arc": [w, z], "value": x_wz},
                    ],
                }
            )
    return violations


def _bipartition_violations(model: OAPCompactModel, mode: int, tol: float) -> list[dict[str, Any]]:
    violations: list[dict[str, Any]] = []
    crossing_index = build_crossing_arc_index(model.points, model.x.keys())

    for (p, q) in model.x.keys():
        x_pq = model.x_relaxed.get((p, q), 0.0)
        src_groups, dst_groups = crossing_index.get((p, q), ({}, {}))

        if mode in {0, 2}:
            for u, arcs in src_groups.items():
                lhs = x_pq + sum(model.x_relaxed.get(arc, 0.0) for arc in arcs)
                if len(arcs) < 1:
                    continue
                rhs = 1.0
                viol = lhs - rhs
                if viol > tol:
                    top_terms = [{"arc": [p, q], "value": x_pq}] + [
                        {"arc": [a, b], "value": model.x_relaxed.get((a, b), 0.0)} for a, b in arcs[:8]
                    ]
                    violations.append(
                        {
                            "family": "bipartition",
                            "mode": "source-fixed",
                            "key_indices": {"p": p, "q": q, "u": u},
                            "lhs": lhs,
                            "rhs": rhs,
                            "violation": viol,
                            "top_terms": top_terms,
                        }
                    )

        if mode in {1, 2}:
            for w, arcs in dst_groups.items():
                lhs = x_pq + sum(model.x_relaxed.get(arc, 0.0) for arc in arcs)
                if len(arcs) < 1:
                    continue
                rhs = 1.0
                viol = lhs - rhs
                if viol > tol:
                    top_terms = [{"arc": [p, q], "value": x_pq}] + [
                        {"arc": [a, b], "value": model.x_relaxed.get((a, b), 0.0)} for a, b in arcs[:8]
                    ]
                    violations.append(
                        {
                            "family": "bipartition",
                            "mode": "destination-fixed",
                            "key_indices": {"p": p, "q": q, "w": w},
                            "lhs": lhs,
                            "rhs": rhs,
                            "violation": viol,
                            "top_terms": top_terms,
                        }
                    )

    return violations


def _triangle_violations(model: OAPCompactModel, mode: int, tol: float) -> list[dict[str, Any]]:
    violations: list[dict[str, Any]] = []
    crossing_index = build_crossing_arc_index(model.points, model.x.keys())

    for i in range(model.N):
        for k in range(model.N):
            if k == i:
                continue
            for p in range(model.N):
                if p == i or p == k:
                    continue
                seg_kp = crossing_index.get((k, p))
                seg_ip = crossing_index.get((i, p))
                seg_ik = crossing_index.get((i, k))
                if seg_kp is None or seg_ip is None or seg_ik is None:
                    continue

                if mode in {0, 2}:
                    arcs_src = seg_kp[0].get(i, []) + seg_ip[0].get(k, []) + seg_ik[0].get(p, [])
                    if len(arcs_src) >= 2:
                        lhs = sum(model.x_relaxed.get(arc, 0.0) for arc in arcs_src)
                        rhs = 1.0
                        viol = lhs - rhs
                        if viol > tol:
                            top_terms = [
                                {"arc": [a, b], "value": model.x_relaxed.get((a, b), 0.0)}
                                for a, b in arcs_src[:10]
                            ]
                            violations.append(
                                {
                                    "family": "triangle_crossing",
                                    "mode": "source-fixed",
                                    "key_indices": {"i": i, "k": k, "p": p},
                                    "lhs": lhs,
                                    "rhs": rhs,
                                    "violation": viol,
                                    "top_terms": top_terms,
                                }
                            )

                if mode in {1, 2}:
                    arcs_dst = seg_kp[1].get(i, []) + seg_ip[1].get(k, []) + seg_ik[1].get(p, [])
                    if len(arcs_dst) >= 2:
                        lhs = sum(model.x_relaxed.get(arc, 0.0) for arc in arcs_dst)
                        rhs = 1.0
                        viol = lhs - rhs
                        if viol > tol:
                            top_terms = [
                                {"arc": [a, b], "value": model.x_relaxed.get((a, b), 0.0)}
                                for a, b in arcs_dst[:10]
                            ]
                            violations.append(
                                {
                                    "family": "triangle_crossing",
                                    "mode": "destination-fixed",
                                    "key_indices": {"i": i, "k": k, "p": p},
                                    "lhs": lhs,
                                    "rhs": rhs,
                                    "violation": viol,
                                    "top_terms": top_terms,
                                }
                            )

    return violations


def _diagnose_combo(model: OAPCompactModel, combo: Combo, top_n: int = 25, tol: float = VIOL_TOL) -> list[dict[str, Any]]:
    all_viol: list[dict[str, Any]] = []
    all_viol.extend(_crossing_violations(model, tol=tol))
    all_viol.extend(_bipartition_violations(model, mode=combo.bip_mode, tol=tol))
    all_viol.extend(_triangle_violations(model, mode=combo.tri_mode, tol=tol))
    all_viol.sort(key=lambda x: float(x["violation"]), reverse=True)
    return all_viol[:top_n]


def _write_excel(df_full: pd.DataFrame, df_best: pd.DataFrame, df_leader: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        df_full.to_excel(writer, sheet_name="full", index=False)
        df_best.to_excel(writer, sheet_name="best_per_instance", index=False)
        df_leader.to_excel(writer, sheet_name="leaderboard", index=False)

        ws = writer.sheets["full"]
        ws_best = writer.sheets["best_per_instance"]
        ws_leader = writer.sheets["leaderboard"]

        fill_header = PatternFill(fill_type="solid", fgColor="D9E1F2")
        fill_best = PatternFill(fill_type="solid", fgColor="375623")
        fill_worst = PatternFill(fill_type="solid", fgColor="9C0006")
        font_header = Font(bold=True)
        font_light = Font(color="FFFFFF", bold=True)
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for worksheet in (ws, ws_best, ws_leader):
            for cell in worksheet[1]:
                cell.fill = fill_header
                cell.font = font_header
                cell.alignment = align

            for col_cells in worksheet.columns:
                col_letter = get_column_letter(col_cells[0].column)
                max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
                worksheet.column_dimensions[col_letter].width = min(max_len + 3, 28)

            worksheet.freeze_panes = "A2"

        headers = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
        best_col = headers.get("IsBest")
        worst_col = headers.get("IsWorst")
        delta_col = headers.get("DeltaVsBaseline")

        if best_col is not None and worst_col is not None:
            for r in range(2, ws.max_row + 1):
                if ws.cell(r, best_col).value:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(r, c).fill = fill_best
                        ws.cell(r, c).font = font_light
                elif ws.cell(r, worst_col).value:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(r, c).fill = fill_worst
                        ws.cell(r, c).font = font_light

        if delta_col is not None and ws.max_row >= 2:
            col_letter = get_column_letter(delta_col)
            rng = f"{col_letter}2:{col_letter}{ws.max_row}"
            ws.conditional_formatting.add(
                rng,
                DataBarRule(start_type="num", start_value=0, end_type="max", end_value=0, color="63C384"),
            )


def main(
    time_limit: int,
    top_k_worst: int,
    max_instances: int | None,
    directions: str,
    export_violations_json: bool,
) -> None:
    if not TSV_PATH.exists():
        print(f"[ERROR] TSV not found: {TSV_PATH}")
        sys.exit(1)

    tsv = pd.read_csv(TSV_PATH, sep="\t")
    tsv = tsv.sort_values(by=["N", "instance"], ascending=[True, True], kind="stable")
    combos = _all_combos()

    if directions == "both":
        dir_flags = [("MIN", False), ("MAX", True)]
    elif directions == "min":
        dir_flags = [("MIN", False)]
    else:
        dir_flags = [("MAX", True)]

    records: list[dict[str, Any]] = []
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}

    processed = 0
    for _, ref in tsv.iterrows():
        inst_name = str(ref["instance"])
        inst_file = INSTANCE_DIR / f"{inst_name}.instance"
        if not inst_file.exists():
            continue

        n = int(ref["N"])
        processed += 1
        if max_instances is not None and processed > max_instances:
            break

        print(f"[{inst_name}] N={n}")

        for direction, maximize in dir_flags:
            per_dir: list[dict[str, Any]] = []

            for combo in combos:
                lp_val, _model = _solve_relaxed(inst_file, maximize=maximize, combo=combo, time_limit=time_limit)
                row = {
                    "Instance": inst_name,
                    "N": n,
                    "Direction": direction,
                    "CrossingV1": int(combo.crossing_on),
                    "BipMode": combo.bip_mode,
                    "TriMode": combo.tri_mode,
                    "ComboID": combo.combo_id,
                    "LP": lp_val,
                }
                per_dir.append(row)
                print(f"  {direction} {combo.combo_id}: LP={lp_val}")

            baseline: float | str = cast(float | str, next((r["LP"] for r in per_dir if r["ComboID"] == "c0_b0_t0"), "-"))
            for row in per_dir:
                lp_here = cast(float | str, row["LP"])
                row["DeltaVsBaseline"] = _bound_delta(direction, lp_here, baseline)

            sorted_rows = sorted(per_dir, key=lambda r: _sort_key(direction, r["LP"]))
            for rank, row in enumerate(sorted_rows, start=1):
                row["Rank"] = rank

            if sorted_rows:
                best_key = sorted_rows[0]["ComboID"]
                worst_key = sorted_rows[-1]["ComboID"]
            else:
                best_key = ""
                worst_key = ""

            for row in per_dir:
                row["IsBest"] = row["ComboID"] == best_key
                row["IsWorst"] = row["ComboID"] == worst_key

            grouped[(inst_name, direction)] = per_dir
            records.extend(per_dir)

    if not records:
        print("No records generated.")
        return

    df_full = pd.DataFrame(records)
    df_best = df_full[df_full["IsBest"]].copy().sort_values(by=["Direction", "Instance"])

    df_leader = (
        df_full[df_full["IsBest"]]
        .groupby(["Direction", "ComboID", "CrossingV1", "BipMode", "TriMode"], as_index=False)
        .size()
        .rename(columns={"size": "Wins"})
        .sort_values(by=["Direction", "Wins"], ascending=[True, False])
    )

    _write_excel(df_full=df_full, df_best=df_best, df_leader=df_leader, output_path=OUTPUT_XLSX)
    print(f"Excel written to: {OUTPUT_XLSX}")

    if export_violations_json:
        violations_report: list[dict[str, Any]] = []
        for (inst_name, direction), rows in grouped.items():
            ordered = sorted(rows, key=lambda r: _sort_key(direction, r["LP"]))
            worst_rows = ordered[-top_k_worst:] if top_k_worst > 0 else []
            instance_file = INSTANCE_DIR / f"{inst_name}.instance"

            for row in worst_rows:
                combo = Combo(
                    crossing_on=bool(cast(int, row["CrossingV1"])),
                    bip_mode=int(cast(int, row["BipMode"])),
                    tri_mode=int(cast(int, row["TriMode"])),
                )
                maximize = direction == "MAX"
                lp_val, model = _solve_relaxed(instance_file, maximize=maximize, combo=combo, time_limit=time_limit)
                top_viol = _diagnose_combo(model=model, combo=combo, top_n=25, tol=VIOL_TOL)
                violations_report.append(
                    {
                        "instance": inst_name,
                        "direction": direction,
                        "combo": {
                            "id": combo.combo_id,
                            "crossing_v1": int(combo.crossing_on),
                            "bip_mode": combo.bip_mode,
                            "triangle_mode": combo.tri_mode,
                        },
                        "lp": lp_val,
                        "rank": int(cast(int, row.get("Rank", -1))),
                        "top_violations": top_viol,
                    }
                )

        OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
        with open(OUTPUT_JSON, "w", encoding="utf-8") as fh:
            json.dump(violations_report, fh, ensure_ascii=False, indent=2)

        print(f"JSON written to: {OUTPUT_JSON}")
    else:
        print("JSON export disabled.")

    print(f"Done. Rows={len(df_full)} | Best rows={len(df_best)}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Compare LP relaxations across crossing/bipartition/triangle combinations.")
    parser.add_argument("--time-limit", type=int, default=TIME_LIMIT, help=f"Time limit per LP solve (default: {TIME_LIMIT}).")
    parser.add_argument("--top-k-worst", type=int, default=2, help="How many worst combos per instance/direction to diagnose.")
    parser.add_argument("--max-instances", type=int, default=None, help="Optional cap for processed instances.")
    parser.add_argument("--directions", choices=["min", "max", "both"], default="both", help="Directions to evaluate.")
    parser.add_argument(
        "--export-violations-json",
        action="store_true",
        default=False,
        help="Export JSON diagnostics for worst combinations (disabled by default).",
    )
    args = parser.parse_args()

    main(
        time_limit=args.time_limit,
        top_k_worst=args.top_k_worst,
        max_instances=args.max_instances,
        directions=args.directions,
        export_violations_json=args.export_violations_json,
    )
