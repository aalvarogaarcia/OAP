"""experiments/solveLP.py
Compute MCF-LP and MCF+Tₖ-LP bounds for every instance that appears in
both ``instance/`` and ``test/data/TablaResultadosA4.tsv``, then write a
formatted Excel comparison.

Run from the repo root:
    .venv/bin/python experiments/solveLP.py

The TSV supplies the reference LP / IP values (computed with SCF subtour).
This script solves the LP relaxation twice per (instance × direction):
  1. MCF subtour only
  2. MCF + T_k lifted-cycle cuts (limited to MAX_TK_ITERATIONS rounds)

For MIN OAP: LP is a lower bound → bigger LP = tighter bound = better.
For MAX OAP: LP is an upper bound → smaller LP = tighter bound = better.

Gap formulae (matching the TSV convention):
  MIN gap % = (IP − LP) / IP × 100
  MAX gap % = (LP − IP) / (CH − IP) × 100
"""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Repo root on sys.path so we can import models/utils from any working dir
# ---------------------------------------------------------------------------
_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT))

import pandas as pd  # noqa: E402 (import after sys.path mutation)
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from models import OAPCompactModel  # noqa: E402
from utils.utils import compute_triangles, read_indexed_instance  # noqa: E402

# ---------------------------------------------------------------------------
# Default configuration (edit here or override via CLI flags)
# ---------------------------------------------------------------------------
INSTANCE_DIR: Path = _REPO_ROOT / "instance"
TSV_PATH: Path = _REPO_ROOT / "test" / "data" / "TablaResultadosA4.tsv"
OUTPUT_PATH: Path = _REPO_ROOT / "outputs" / "CSV" / "lp_comparison.xlsx"

OBJECTIVE: str = "Fekete"   # area objective (same IP optimum as "External")
MODE: int = 0
SUM_CONSTRAIN: bool = True
STRENGTHEN: bool = False     # matches main.py interactive default

MAX_TK_ITERATIONS: int = None  # LP cutting-plane iterations for T_k cuts
TIME_LIMIT: int = 300       # seconds per LP solve

# ---------------------------------------------------------------------------
# Sentinel returned when solve fails / no solution
# ---------------------------------------------------------------------------
_NA = "-"


# ---------------------------------------------------------------------------
# LP solve helper
# ---------------------------------------------------------------------------

def _solve_lp(
    instance_file: Path,
    maximize: bool,
    use_tk_cuts: bool,
    time_limit: int,
) -> float | str:
    """Build a fresh OAPCompactModel, solve its LP relaxation, return the
    LP objective value (Shoelace area).

    Returns ``"-"`` on failure (infeasible, timeout without solution, etc.).
    """
    direction = "max" if maximize else "min"
    variant = "tk" if use_tk_cuts else "mcf"
    name = f"{instance_file.stem}_{direction}_{variant}"

    points = read_indexed_instance(str(instance_file))
    triangles = compute_triangles(points)
    model = OAPCompactModel(points, triangles, name=name)
    model.build(
        objective=OBJECTIVE,  # type: ignore[arg-type]
        mode=MODE,
        maximize=maximize,
        subtour="MCF",
        sum_constrain=SUM_CONSTRAIN,
        strengthen=STRENGTHEN,
        use_tk_cuts=use_tk_cuts,
    )
    model.solve(
        time_limit=time_limit,
        relaxed=True,
    )
    return model.get_objval_lp()


# ---------------------------------------------------------------------------
# Gap & comparison helpers
# ---------------------------------------------------------------------------

def _gap_min(lp: float | str, ip: float) -> float | str:
    """MIN gap: (IP − LP) / IP × 100."""
    if not isinstance(lp, (int, float)) or ip == 0:
        return _NA
    return (ip - lp) / ip * 100


def _gap_max(lp: float | str, ip: float, ch: float) -> float | str:
    """MAX gap: (LP − IP) / (CH − IP) × 100."""
    if not isinstance(lp, (int, float)):
        return _NA
    denom = ch - ip
    if denom == 0:
        return _NA
    return (lp - ip) / denom * 100


def _better_min(lp_new: float | str, lp_ref: float) -> str:
    """Return '✓' when the new MIN LP bound is strictly tighter (larger)."""
    if not isinstance(lp_new, (int, float)):
        return _NA
    return "✓" if lp_new > lp_ref * (1 + 1e-6) else ""


def _better_max(lp_new: float | str, lp_ref: float) -> str:
    """Return '✓' when the new MAX LP bound is strictly tighter (smaller)."""
    if not isinstance(lp_new, (int, float)):
        return _NA
    return "✓" if lp_new < lp_ref * (1 - 1e-6) else ""


# ---------------------------------------------------------------------------
# Excel formatting helpers
# ---------------------------------------------------------------------------

_FILL_INFO = PatternFill(fill_type="solid", fgColor="BDD7EE")    # blue  – identity/ref
_FILL_MCF = PatternFill(fill_type="solid", fgColor="E2EFDA")     # green – MCF
_FILL_TK = PatternFill(fill_type="solid", fgColor="FCE4D6")      # orange – MCF+Tk
_FILL_BETTER = PatternFill(fill_type="solid", fgColor="375623")  # dark green – better flag
_FONT_HEADER = Font(bold=True)
_FONT_BETTER = Font(bold=True, color="FFFFFF")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_FMT_NUM = "#,##0.00"
_FMT_PCT = '0.00"%"'


def _col_fill(col_name: str) -> PatternFill:
    if "Better" in col_name or col_name in ("Instance", "N", "Conv.Hull(N)"):
        return _FILL_INFO
    if "Tk" in col_name or "+Tk" in col_name:
        return _FILL_TK
    if "MCF" in col_name:
        return _FILL_MCF
    return _FILL_INFO


def _write_xlsx(df: pd.DataFrame, path: Path) -> None:
    """Write *df* to *path* with colour-coded headers and formatting."""
    path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(str(path), engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="LP Comparison", index=False)
        ws = writer.sheets["LP Comparison"]

        # --- Header row formatting ---
        for cell in ws[1]:
            col_name = str(cell.value or "")
            cell.font = _FONT_HEADER
            cell.fill = _col_fill(col_name)
            cell.alignment = _ALIGN_CENTER

        # --- Identify column indices by header name ---
        headers: dict[str, int] = {
            str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)
        }
        better_cols: set[int] = {v for k, v in headers.items() if "Better" in k}
        gap_cols: set[int] = {v for k, v in headers.items() if "Gap" in k}
        num_cols: set[int] = {
            v for k, v in headers.items()
            if any(tok in k for tok in ("LP", "IP", "Conv.Hull"))
        }

        # --- Data rows ---
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                c = cell.column
                val = cell.value
                if c in better_cols:
                    cell.alignment = _ALIGN_CENTER
                    if val == "✓":
                        cell.fill = _FILL_BETTER
                        cell.font = _FONT_BETTER
                elif c in gap_cols and isinstance(val, (int, float)):
                    cell.number_format = _FMT_PCT
                elif c in num_cols and isinstance(val, (int, float)):
                    cell.number_format = _FMT_NUM

        # --- Auto column widths ---
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0) for c in col_cells
            )
            ws.column_dimensions[col_letter].width = min(max_len + 3, 22)

        # --- Freeze header row ---
        ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(time_limit: int = TIME_LIMIT) -> None:  # noqa: C901
    if not TSV_PATH.exists():
        print(f"[ERROR] TSV not found: {TSV_PATH}")
        sys.exit(1)

    tsv = pd.read_csv(TSV_PATH, sep="\t")
    print(f"TSV loaded: {len(tsv)} instances.\n")

    records: list[dict[str, Any]] = []

    for _, ref in tsv.iterrows():
        inst_name = str(ref["instance"])
        inst_file = INSTANCE_DIR / f"{inst_name}.instance"

        if not inst_file.exists():
            print(f"[SKIP] {inst_name}: instance file not found")
            continue

        ch = float(ref["conv.hull(N)"])
        min_ip = float(ref["MIN_IPvalue"])
        max_ip = float(ref["MAX_IPvalue"])
        lp_min_ref = float(ref["MIN_LPvalue"])
        lp_max_ref = float(ref["MAX_LPvalue"])
        gap_min_ref = float(ref["MIN_LPgap"])
        gap_max_ref = float(ref["MAX_LPgap"])

        n = int(ref["N"])
        print(f"[{inst_name}]  N={n}")

        # Solve 4 LP relaxations
        print(f"  MIN MCF ...", end="  ", flush=True)
        lp_min_mcf = _solve_lp(inst_file, maximize=False, use_tk_cuts=False, time_limit=time_limit)
        print(f"LP={lp_min_mcf}")

        print(f"  MIN MCF+Tk ...", end="  ", flush=True)
        lp_min_tk = _solve_lp(inst_file, maximize=False, use_tk_cuts=True, time_limit=time_limit)
        print(f"LP={lp_min_tk}")

        print(f"  MAX MCF ...", end="  ", flush=True)
        lp_max_mcf = _solve_lp(inst_file, maximize=True, use_tk_cuts=False, time_limit=time_limit)
        print(f"LP={lp_max_mcf}")

        print(f"  MAX MCF+Tk ...", end="  ", flush=True)
        lp_max_tk = _solve_lp(inst_file, maximize=True, use_tk_cuts=True, time_limit=time_limit)
        print(f"LP={lp_max_tk}")

        records.append(
            {
                # --- Identity ---
                "Instance": inst_name,
                "N": n,
                "Conv.Hull(N)": ch,
                # --- MIN: reference ---
                "MIN LP (ref)": lp_min_ref,
                "MIN Gap ref (%)": gap_min_ref,
                "MIN IP": min_ip,
                # --- MIN: MCF ---
                "MIN LP MCF": lp_min_mcf,
                "MIN Gap MCF (%)": _gap_min(lp_min_mcf, min_ip),
                "MIN MCF Better": _better_min(lp_min_mcf, lp_min_ref),
                # --- MIN: MCF+Tk ---
                "MIN LP MCF+Tk": lp_min_tk,
                "MIN Gap MCF+Tk (%)": _gap_min(lp_min_tk, min_ip),
                "MIN Tk Better": _better_min(lp_min_tk, lp_min_ref),
                # --- MAX: reference ---
                "MAX LP (ref)": lp_max_ref,
                "MAX Gap ref (%)": gap_max_ref,
                "MAX IP": max_ip,
                # --- MAX: MCF ---
                "MAX LP MCF": lp_max_mcf,
                "MAX Gap MCF (%)": _gap_max(lp_max_mcf, max_ip, ch),
                "MAX MCF Better": _better_max(lp_max_mcf, lp_max_ref),
                # --- MAX: MCF+Tk ---
                "MAX LP MCF+Tk": lp_max_tk,
                "MAX Gap MCF+Tk (%)": _gap_max(lp_max_tk, max_ip, ch),
                "MAX Tk Better": _better_max(lp_max_tk, lp_max_ref),
            }
        )

    if not records:
        print("\nNo instances processed — nothing to write.")
        return

    df = pd.DataFrame(records)
    _write_xlsx(df, OUTPUT_PATH)
    print(f"\n{'=' * 60}")
    print(f"Done. {len(records)} instances processed.")
    print(f"Output: {OUTPUT_PATH}")
    print(f"{'=' * 60}")

    # Quick terminal summary
    _print_summary(df)


def _print_summary(df: pd.DataFrame) -> None:
    """Print a compact summary table to stdout."""
    print("\nSummary (✓ = new LP tighter than reference):\n")
    header = (
        f"{'Instance':<28} {'N':>3}  "
        f"{'MIN MCF':>9} {'MIN Tk':>9}  "
        f"{'MAX MCF':>9} {'MAX Tk':>9}"
    )
    print(header)
    print("-" * len(header))
    for _, r in df.iterrows():
        flag = lambda col: "✓" if r.get(col) == "✓" else " "  # noqa: E731
        print(
            f"{str(r['Instance']):<28} {int(r['N']):>3}  "
            f"{flag('MIN MCF Better'):>9} {flag('MIN Tk Better'):>9}  "
            f"{flag('MAX MCF Better'):>9} {flag('MAX Tk Better'):>9}"
        )


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Solve LP relaxations and compare with TSV reference.")
    parser.add_argument(
        "--time-limit",
        type=int,
        default=TIME_LIMIT,
        metavar="SECONDS",
        help=f"Time limit per LP solve (default: {TIME_LIMIT}s).",
    )
    args = parser.parse_args()
    main(time_limit=args.time_limit)
